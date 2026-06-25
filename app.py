from flask import Flask, render_template, request, redirect, url_for, session, send_file
import json
import os
import uuid
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from io import BytesIO
import gspread

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# ==========================================
# ⚙️ 시스템 기본 설정 (부서/예산/계정 관리)
# ==========================================

TEAM_BUDGETS = {
    "시운전팀": 1000000,
    "생산팀": 500000,
    "판금생산팀" : 0,
    "생산설계팀": 0,
    "영업팀": 500000,
    "영업2팀": 0,
    "영업3팀": 0,
    "전장팀": 800000,
    "법카2536": 0,
    "법카6035": 0,
    "법카7547": 0,
    "법카0624": 0
}
TEAMS_LIST = list(TEAM_BUDGETS.keys())

USER_CREDENTIALS = {
    "admin": {"password": "01234", "name": "관리자", "team": "관리자"},
    "시운전": {"password": "1234", "name": "시운전", "team": "시운전팀"},
    "생산": {"password": "1234", "name": "생산", "team": "생산팀"},
    "판금": {"password": "1234", "name": "판금생산", "team": "판금생산팀"},
    "생산설계": {"password": "1234", "name": "생산설계", "team": "생산설계팀"},
    "영업": {"password": "1234", "name": "영업", "team": "영업팀"},
    "영업2": {"password": "1234", "name": "영업2", "team": "영업2팀"},
    "영업3": {"password": "1234", "name": "영업3", "team": "영업3팀"},
    "전장": {"password": "1234", "name": "전장", "team": "전장팀"},
    "법카2536": {"password": "1234", "name": "법카2536", "team": "법카2536"},
    "법카6035": {"password": "1234", "name": "법카6035", "team": "법카6035"},
    "법카7547": {"password": "1234", "name": "법카7547", "team": "법카7547"},
    "법카0624": {"password": "1234", "name": "법카0624", "team": "법카0624"}
}

CATEGORIES = ["교통비", "주차비", "식비", "숙박비", "소모품비", "차량유지비", "운반비", "기타"]

# ==========================================
# 🌟 구글 스프레드시트 DB 연동 설정 🌟
# ==========================================
SHEET_URL = "https://docs.google.com/spreadsheets/d/1wJrlVE1RfDR48T4IliC2xjsvHXC-6gpWUZBeCqUxflE/edit?gid=0#gid=0"

try:
    gc = gspread.service_account(filename='credentials.json')
    doc = gc.open_by_url(SHEET_URL)
    ws = doc.sheet1
    print("✅ 구글 스프레드시트 연결 성공!")
except Exception as e:
    print("❌ 구글 시트 연결 실패! credentials.json 파일과 시트 공유 상태를 확인하세요:", e)
    ws = None

HEADERS = ["trip_id", "order", "team", "date", "user", "place", "content", "items_desc", "total_amount", "details_json"]

def get_all_trips():
    if not ws: return []
    try:
        records = ws.get_all_records()
        for r in records:
            r['order'] = int(r.get('order') if r.get('order') else 999)
            r['total_amount'] = int(r.get('total_amount') if r.get('total_amount') else 0)
            r['trip_id'] = str(r.get('trip_id', ''))
        return sorted(records, key=lambda x: x['order'])
    except Exception as e:
        print("데이터 로드 에러:", e)
        return []

def save_all_trips(trips_list):
    if not ws: return
    try:
        # 1. 데이터 준비
        values = [HEADERS] + [[str(t.get(h, "")) for h in HEADERS] for t in trips_list]
        
        # 2. 여유분 추가 (선택 사항이지만 기존 로직 유지)
        empty_row = [""] * len(HEADERS)
        values.extend([empty_row] * 50)
        
        # 3. 🚨 수정된 부분: 키워드 인자 제거하고 직접 전달 🚨
        # range_name을 첫 번째 인자로, 데이터를 두 번째 인자로 전달합니다.
        ws.update(range_name="A1", values=values)
        
    except Exception as e:
        print("데이터 저장 실패:", e)
        
# ==========================================
# 라우팅 (페이지 기능)
# ==========================================

@app.route('/')
def login_page():
    if 'username' in session: return redirect(url_for('index'))
    html = """
    <!DOCTYPE html><html><head><meta charset="UTF-8"><title>경비 정산 로그인</title>
    <style>
        body { font-family: '맑은 고딕', sans-serif; background: #f0f4f8; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; }
        .login-box { background: white; padding: 40px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); text-align: center; width: 300px; }
        input { width: 100%; padding: 12px; margin: 10px 0; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box; }
        button { width: 100%; padding: 12px; background: #1e3a8a; color: white; border: none; border-radius: 4px; font-weight: bold; cursor: pointer; font-size: 15px; }
    </style>
    </head><body>
    <div class="login-box">
        <h2 style="color:#1e3a8a; margin-bottom:20px;">경비 정산 시스템</h2>
        <form action="/login" method="post">
            <input type="text" name="username" placeholder="아이디 (예: admin, 시운전)" required>
            <input type="password" name="password" placeholder="비밀번호" required>
            <button type="submit">로그인</button>
        </form>
    </div>
    </body></html>
    """
    return html

@app.route('/login', methods=['POST'])
def do_login():
    uid = request.form.get('username')
    upass = request.form.get('password')
    if uid in USER_CREDENTIALS and USER_CREDENTIALS[uid]['password'] == upass:
        session['user_id'] = uid
        session['username'] = USER_CREDENTIALS[uid]['name']
        session['team'] = USER_CREDENTIALS[uid]['team']
        return redirect(url_for('index'))
    return "<script>alert('아이디 또는 비밀번호가 올바르지 않습니다.'); history.back();</script>"

@app.route('/index')
def index():
    if 'username' not in session: return redirect(url_for('login_page'))
    username = session.get('username', '게스트')
    team = session.get('team', '시운전팀')
    
    current_month = request.args.get('search_month', datetime.now().strftime('%Y-%m'))
    month_start_date = f"{current_month}-01"
    month_end_date = f"{current_month}-31"
    
    ALL_TRIPS = get_all_trips()
    filtered_trips = [t for t in ALL_TRIPS if str(t.get('date', '')).startswith(current_month)]
    
    raw_stats_list = []
    
    dashboard_stats = {'총합': 0}
    for t_name in TEAMS_LIST:
        dashboard_stats[t_name] = 0
    
    for t in ALL_TRIPS:
        try: details = json.loads(t.get('details_json', '[]'))
        except: details = []
        for item in details:
            stat_item = {
                "date": t.get('date', ''), "team": t.get('team', ''),
                "place": t.get('place', ''), "user": t.get('user', '알수없음'), 
                "category": item.get('category', '기타'), "amount": int(item.get('amount', 0))
            }
            raw_stats_list.append(stat_item)
            
            if str(t.get('date', '')).startswith(current_month):
                amt = int(item.get('amount', 0))
                dashboard_stats['총합'] += amt
                
                # 📌 치명적 오류 수정: 무조건 '팀'을 붙이던 로직에서 이름 그대로 매칭되도록 변경
                raw_team = str(t.get('team', '')).strip()
                std_team = raw_team
                if std_team not in TEAMS_LIST:
                    if std_team + '팀' in TEAMS_LIST:
                        std_team += '팀'
                
                if std_team in dashboard_stats:
                    dashboard_stats[std_team] += amt

    return render_template('index.html', username=username, team=team, current_month=current_month,
        month_start_date=month_start_date, month_end_date=month_end_date, trips=filtered_trips,
        categories=CATEGORIES, dashboard_stats=dashboard_stats, raw_stats_json=json.dumps(raw_stats_list, ensure_ascii=False),
        teams=TEAMS_LIST)

@app.route('/expense/add', methods=['POST'])
def add_expense():
    expense_date = request.form.get('expense_date') 
    user_name = request.form.get('user_name')       
    place = request.form.get('place')
    content = request.form.get('content')
    search_month = request.form.get('search_month', datetime.now().strftime('%Y-%m'))
    
    user_team = session.get('team', TEAMS_LIST[0])
    if user_team == "관리자" and request.form.get('target_team'):
        user_team = request.form.get('target_team')
        
    receipt_cats = request.form.getlist('receipt_category')
    receipt_amts = request.form.getlist('receipt_amount')
    
    details, total_amount, desc_parts = [], 0, []
    for i in range(len(receipt_cats)):
        cat = receipt_cats[i]
        try: amt = int(receipt_amts[i]) if receipt_amts[i] else 0
        except: amt = 0
        if cat and amt > 0:
            details.append({"id": f"r_{uuid.uuid4().hex[:6]}", "category": cat, "amount": amt})
            total_amount += amt
            desc_parts.append(f"{cat}: {amt:,}원")
            
    items_desc = " | ".join(desc_parts) if desc_parts else "등록된 영수증 없음"
    
    ALL_TRIPS = get_all_trips()
    
    new_trip = {
        "trip_id": str(uuid.uuid4().hex[:8]), "order": len(ALL_TRIPS) + 1,
        "team": user_team, "date": expense_date, "user": user_name,
        "place": place, "content": content, "items_desc": items_desc,
        "total_amount": total_amount, "details_json": json.dumps(details, ensure_ascii=False)
    }
    
    ALL_TRIPS.append(new_trip)
    save_all_trips(ALL_TRIPS) 
        
    return redirect(url_for('index', search_month=search_month))

@app.route('/expense/edit_submit', methods=['POST'])
def edit_submit():
    trip_id = request.form.get('trip_id')
    search_month = request.form.get('search_month')
    sub_ids = request.form.getlist('sub_receipt_ids')
    sub_categories = request.form.getlist('sub_receipt_categories')
    sub_amounts = request.form.getlist('sub_receipt_amounts')
    
    ALL_TRIPS = get_all_trips()
    for t in ALL_TRIPS:
        if str(t.get('trip_id')) == str(trip_id):
            t['date'] = request.form.get('date')
            t['user'] = request.form.get('user')
            t['place'] = request.form.get('place')
            t['content'] = request.form.get('content')
            
            new_details, total_amount, desc_parts = [], 0, []
            for i in range(len(sub_ids)):
                try: amt = int(sub_amounts[i])
                except: amt = 0
                cat = sub_categories[i]
                new_details.append({"id": sub_ids[i], "category": cat, "amount": amt})
                total_amount += amt
                desc_parts.append(f"{cat}: {amt:,}원")
                
            t['total_amount'] = total_amount
            t['items_desc'] = " | ".join(desc_parts)
            t['details_json'] = json.dumps(new_details, ensure_ascii=False)
            break
            
    save_all_trips(ALL_TRIPS)
    return redirect(url_for('index', search_month=search_month))

@app.route('/expense/reorder', methods=['POST'])
def reorder():
    trip_ids = request.form.getlist('trip_ids')
    search_month = request.form.get('search_month')
    ALL_TRIPS = get_all_trips()
    
    for index, tid in enumerate(trip_ids):
        for t in ALL_TRIPS:
            if str(t.get('trip_id')) == str(tid):
                t['order'] = index + 1
                break
                
    save_all_trips(ALL_TRIPS)
    return redirect(url_for('index', search_month=search_month))

@app.route('/expense/delete/<trip_id>')
def delete_expense(trip_id):
    search_month = request.args.get('search_month', datetime.now().strftime('%Y-%m'))
    ALL_TRIPS = get_all_trips()
    ALL_TRIPS = [t for t in ALL_TRIPS if str(t.get('trip_id')) != str(trip_id)]
    
    save_all_trips(ALL_TRIPS)
    return redirect(url_for('index', search_month=search_month))

@app.route('/download/cover')
def download_cover():
    target_team = request.args.get('team', 'ALL')
    target_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    ALL_TRIPS = get_all_trips()
    
    if target_team == 'ALL':
        raw_data = [t for t in ALL_TRIPS if str(t.get('date', '')).startswith(target_month)]
        display_team_title = "전사 통합"
    else:
        raw_data = [t for t in ALL_TRIPS if is_match_team(str(t.get('team', '')), target_team) and str(t.get('date', '')).startswith(target_month)]
        display_team_title = target_team

    raw_data.sort(key=lambda x: (int(x.get('order', 999)), str(x.get('date', ''))))
    wb = openpyxl.Workbook()
    
    font_title = Font(name='맑은 고딕', size=18, bold=True, color='000080')
    font_header = Font(name='맑은 고딕', size=11, bold=True)
    font_main = Font(name='맑은 고딕', size=10)
    font_sum = Font(name='맑은 고딕', size=11, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_header = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    fill_sum = PatternFill(start_color='EBF5FF', end_color='EBF5FF', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
    align_right = Alignment(horizontal='right', vertical='center', shrink_to_fit=True)
    align_left = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)

    ws1 = wb.active
    ws1.title = f"{target_month[5:7]}월 정산서"
    ws1.merge_cells('A1:E2')
    ws1['A1'] = f"{target_month[5:7]}월 경비 사용내역서"
    ws1['A1'].font = font_title; ws1['A1'].alignment = align_left

    display_categories = ["교통비", "식비", "숙박비", "소모품비", "차량유지비", "기타"]
    headers1 = ["순번", "일자", "내 용", "출장지", "금액(합계)"] + display_categories + ["사용자"]

    approve_headers = ["작성", "검토", "검토", "승인"]
    for i, h in enumerate(approve_headers):
        col_idx = 9 + i 
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = font_header; cell.alignment = align_center; cell.border = thin_border
        cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        ws1.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
        for r in range(2, 4): ws1.cell(row=r, column=col_idx).border = thin_border
        
    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 24
    ws1.row_dimensions[3].height = 24

    ws1.merge_cells('A4:E4')
    ws1['A4'] = f"작성일자: {datetime.now().strftime('%Y년 %m월 %d일')}  /  부서: {display_team_title}"
    ws1['A4'].font = font_main
    ws1.row_dimensions[4].height = 24

    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=5, column=col_idx, value=h)
        cell.font = font_header; cell.alignment = align_center; cell.border = thin_border; cell.fill = fill_header
    ws1.row_dimensions[5].height = 32

    r_idx = 6
    for idx, trip in enumerate(raw_data, 1):
        ws1.cell(row=r_idx, column=1, value=idx).alignment = align_center
        raw_date = str(trip.get('date', ''))
        ws1.cell(row=r_idx, column=2, value=raw_date[-2:] if len(raw_date)>=10 else raw_date).alignment = align_center
        
        display_content = f"[{trip['team']}] {trip['content']}" if target_team == 'ALL' else trip.get('content', '')
        ws1.cell(row=r_idx, column=3, value=display_content).alignment = align_left
        ws1.cell(row=r_idx, column=4, value=trip.get('place', '')).alignment = align_center
        
        t_cell = ws1.cell(row=r_idx, column=5, value=int(trip.get('total_amount', 0)))
        t_cell.font = Font(name='맑은 고딕', size=10, bold=True); t_cell.number_format = '#,##0'; t_cell.alignment = align_right
        
        cat_sums = {c: 0 for c in display_categories}
        try:
            details = json.loads(trip.get('details_json', '[]'))
            for item in details:
                c_name = item.get('category', '기타')
                amt = int(item.get('amount', 0))
                if c_name in ['교통비', '주차비']: cat_sums['교통비'] += amt
                elif c_name in ['운반비', '기타']: cat_sums['기타'] += amt
                elif c_name in cat_sums: cat_sums[c_name] += amt
                else: cat_sums['기타'] += amt
        except: pass
        
        for c_idx, cat_name in enumerate(display_categories, 6):
            v_cell = ws1.cell(row=r_idx, column=c_idx)
            v_cell.value = cat_sums[cat_name] if cat_sums[cat_name] > 0 else ""
            v_cell.number_format = '#,##0'; v_cell.alignment = align_right
            
        ws1.cell(row=r_idx, column=12, value=trip.get('user', '')).alignment = align_center
        
        for c in range(1, len(headers1)+1):
            cell = ws1.cell(row=r_idx, column=c)
            if c != 5: cell.font = font_main
            cell.border = thin_border
            
        ws1.row_dimensions[r_idx].height = 28
        r_idx += 1

    sum_row_idx = r_idx
    ws1.merge_cells(start_row=sum_row_idx, start_column=1, end_row=sum_row_idx, end_column=4)
    ws1.cell(row=sum_row_idx, column=1, value="합   계").font = font_sum
    ws1.cell(row=sum_row_idx, column=1).alignment = align_center
    
    for c in range(1, len(headers1)+1):
        ws1.cell(row=sum_row_idx, column=c).border = thin_border
        ws1.cell(row=sum_row_idx, column=c).fill = fill_sum
    
    for c in range(5, 12):
        col_letter = openpyxl.utils.get_column_letter(c)
        sum_cell = ws1.cell(row=sum_row_idx, column=c, value=f"=SUM({col_letter}6:{col_letter}{sum_row_idx-1})")
        sum_cell.font = font_sum; sum_cell.number_format = '#,##0'; sum_cell.alignment = align_right
        
    ws1.row_dimensions[sum_row_idx].height = 30

    team_budget = TEAM_BUDGETS.get(display_team_title, 0) if target_team != 'ALL' else 0
    budget_str = f"{team_budget:,.0f}" if team_budget > 0 else "0"
    
    r_idx += 2
    ws1.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=12)
    summary_cell = ws1.cell(row=r_idx, column=1)
    
    if team_budget > 0:
        summary_cell.value = f'="가지급금금액(이월잔액포함) [ {budget_str} ]   -   총경비사용금액 [ " & TEXT(E{sum_row_idx}, "#,##0") & " ]   =   잔액 [ " & TEXT({team_budget}-E{sum_row_idx}, "#,##0") & " ]"'
    else:
        summary_cell.value = f'="전체 통합 경비 합계액 [ " & TEXT(E{sum_row_idx}, "#,##0") & " ] 원"'
        
    summary_cell.font = Font(name='맑은 고딕', size=12, bold=True, color='1F2937')
    summary_cell.alignment = align_center
    ws1.row_dimensions[r_idx].height = 36

    widths1 = {1: 4, 2: 4, 3: 40, 4: 5} 
    for i in range(5, 13): widths1[i] = 9 
    for col_idx, w in widths1.items():
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    ws2 = wb.create_sheet(title="상세내역")
    ws2.merge_cells('A1:C2')
    ws2['A1'] = "지출 항목별 상세 증빙내역"
    ws2['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='374151')
    ws2['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    headers2 = ["순번", "사용일자", "부서명", "사용자", "경비구분", "지출 내용 및 세부 목적", "출장지", "사용 금액", "비고"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = font_header; cell.alignment = align_center; cell.border = thin_border; cell.fill = fill_header
        
    d_idx = 5
    detail_count = 1
    for trip in raw_data:
        try:
            details = json.loads(trip.get('details_json', '[]'))
            for item in details:
                ws2.cell(row=d_idx, column=1, value=detail_count).alignment = align_center
                ws2.cell(row=d_idx, column=2, value=str(trip.get('date', ''))[5:]).alignment = align_center
                ws2.cell(row=d_idx, column=3, value=trip.get('team', '')).alignment = align_center
                ws2.cell(row=d_idx, column=4, value=trip.get('user', '')).alignment = align_center
                ws2.cell(row=d_idx, column=5, value=item.get('category', '')).alignment = align_center
                ws2.cell(row=d_idx, column=6, value=trip.get('content', '')).alignment = align_left
                ws2.cell(row=d_idx, column=7, value=trip.get('place', '')).alignment = align_center
                
                amt_cell = ws2.cell(row=d_idx, column=8, value=int(item.get('amount', 0)))
                amt_cell.number_format = '#,##0'; amt_cell.alignment = align_right
                ws2.cell(row=d_idx, column=9, value="확인완료").alignment = align_center
                
                for c in range(1, 10):
                    cell = ws2.cell(row=d_idx, column=c)
                    cell.border = thin_border; cell.font = font_main
                d_idx += 1; detail_count += 1
        except: pass

    widths2 = [5, 11, 12, 10, 12, 35, 15, 14, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"정산서_{target_team}_{target_month}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
